import io
from pathlib import Path

import openpyxl
import pytest

from app.parsers import xlsx_parser

FIXTURES = Path(__file__).parent.parent / "fixtures"


def _workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_happy_path_single_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["order_id", "customer_name", "total"])
    ws.append([1, "Acme Corp", 100.0])
    ws.append([2, "Beta LLC", 250.5])
    table = xlsx_parser.parse(_workbook_bytes(wb), "orders.xlsx")
    assert table.columns == ["order_id", "customer_name", "total"]
    assert len(table.rows) == 2
    assert table.rows[0]["customer_name"] == "Acme Corp"


def test_picks_largest_of_multiple_sheets_and_warns():
    content = (FIXTURES / "two_sheet.xlsx").read_bytes()
    table = xlsx_parser.parse(content, "two_sheet.xlsx")
    assert table.columns == ["order_id", "customer_name", "total"]
    assert len(table.rows) == 8
    assert any("Orders" in w and "Summary" in w for w in table.warnings)


def test_header_row_not_at_top_is_detected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Company Confidential Export"])
    ws.append(["Generated 2026-01-01"])
    ws.append(["order_id", "customer_name", "total"])
    ws.append([1, "Acme Corp", 100.0])
    ws.append([2, "Beta LLC", 250.5])

    table = xlsx_parser.parse(_workbook_bytes(wb), "titled.xlsx")
    assert table.columns == ["order_id", "customer_name", "total"]
    assert len(table.rows) == 2
    assert any("header row" in w.lower() for w in table.warnings)


def test_merged_cells_propagate_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["region", "order_id", "total"])
    ws.append(["West", 1, 100])
    ws.append(["West", 2, 200])
    ws.merge_cells("A2:A3")

    table = xlsx_parser.parse(_workbook_bytes(wb), "merged.xlsx")
    assert len(table.rows) == 2
    assert table.rows[0]["region"] == "West"
    assert table.rows[1]["region"] == "West"


def test_empty_workbook_returns_warning_not_crash():
    wb = openpyxl.Workbook()
    table = xlsx_parser.parse(_workbook_bytes(wb), "empty.xlsx")
    assert table.rows == []
    assert table.warnings


def test_malformed_bytes_do_not_crash():
    table = xlsx_parser.parse(b"not a real workbook", "broken.xlsx")
    assert table.rows == []
    assert table.warnings
    assert "Failed to open workbook" in table.warnings[0]
