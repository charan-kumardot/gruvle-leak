"""
Excel parsing (.xlsx via openpyxl, legacy .xls via xlrd).

Handles:
  - multiple sheets: picks the sheet with the most data rows, warns about
    the sheet chosen and which others were skipped
  - merged cells: the value in a merged region is propagated to every cell
    in that region before header/row detection runs
  - header rows that aren't row 0: scans the first 10 rows and picks the
    one that looks most like a header (highest fraction of non-numeric,
    non-null, unique string cells)
"""
from __future__ import annotations

from typing import Any

import pandas as pd

from app.parsers._common import dataframe_to_table, dedupe_columns
from app.parsers.base import ParsedTable

_MAX_HEADER_SCAN_ROWS = 10


def _fill_merged_ranges(grid: list[list[Any]], ranges: list[tuple[int, int, int, int]]) -> None:
    """
    ranges: list of (row_start, row_end_exclusive, col_start, col_end_exclusive),
    0-indexed. Mutates grid in place, propagating the top-left value of each
    merged range across the whole range.
    """
    for r1, r2, c1, c2 in ranges:
        if r1 >= len(grid) or c1 >= len(grid[r1]):
            continue
        top_value = grid[r1][c1]
        for r in range(r1, min(r2, len(grid))):
            row = grid[r]
            for c in range(c1, min(c2, len(row))):
                if row[c] in (None, ""):
                    row[c] = top_value


def _read_xlsx_sheets(content: bytes) -> dict[str, list[list[Any]]]:
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    sheets: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        merged_ranges = []
        for merged in ws.merged_cells.ranges:
            merged_ranges.append((merged.min_row - 1, merged.max_row, merged.min_col - 1, merged.max_col))
        _fill_merged_ranges(grid, merged_ranges)
        sheets[ws.title] = grid
    return sheets


def _read_xls_sheets(content: bytes) -> dict[str, list[list[Any]]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet in book.sheets():
        grid = [
            [sheet.cell_value(r, c) if c < sheet.ncols else None for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        # Normalize empty-string cells (xlrd default for blanks) to None.
        for row in grid:
            for i, v in enumerate(row):
                if v == "":
                    row[i] = None
        merged_ranges = [(rlo, rhi, clo, chi) for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", [])]
        _fill_merged_ranges(grid, merged_ranges)
        sheets[sheet.name] = grid
    return sheets


def _non_empty_row_count(grid: list[list[Any]]) -> int:
    count = 0
    for row in grid:
        if any(v is not None and str(v).strip() != "" for v in row):
            count += 1
    return count


def _is_numeric_like(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


def _score_header_row(row: list[Any]) -> float:
    if not row:
        return -1.0
    non_null = [v for v in row if v is not None and str(v).strip() != ""]
    if not non_null:
        return -1.0
    non_numeric = [v for v in non_null if not _is_numeric_like(v)]
    unique_strs = {str(v).strip().lower() for v in non_numeric}

    frac_filled = len(non_null) / len(row)
    frac_non_numeric = len(non_numeric) / len(non_null)
    frac_unique = (len(unique_strs) / len(non_numeric)) if non_numeric else 0.0
    return (frac_filled * 0.3) + (frac_non_numeric * 0.4) + (frac_unique * 0.3)


def _find_header_row(grid: list[list[Any]]) -> int:
    scan_limit = min(_MAX_HEADER_SCAN_ROWS, len(grid))
    best_idx, best_score = 0, -1.0
    for i in range(scan_limit):
        score = _score_header_row(grid[i])
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _grid_to_table(grid: list[list[Any]], warnings: list[str]) -> ParsedTable:
    if not grid or _non_empty_row_count(grid) == 0:
        return ParsedTable(columns=[], rows=[], warnings=warnings + ["Sheet contains no data."])

    header_idx = _find_header_row(grid)
    if header_idx != 0:
        warnings.append(f"Detected header row at row {header_idx + 1} (rows above it appear to be a title/preamble).")

    header = grid[header_idx]
    width = len(header)
    raw_names = [("" if v is None else str(v).strip()) for v in header]
    columns = dedupe_columns(raw_names)

    data_rows = grid[header_idx + 1:]
    df = pd.DataFrame(data_rows, columns=columns)
    if df.shape[1] != width:
        # Defensive: pandas should already align widths from a rectangular grid.
        pass

    return dataframe_to_table(df, warnings)


def parse(content: bytes, filename: str) -> ParsedTable:
    warnings: list[str] = []
    lower = filename.lower()

    try:
        if lower.endswith(".xls") and not lower.endswith(".xlsx"):
            sheets = _read_xls_sheets(content)
        else:
            sheets = _read_xlsx_sheets(content)
    except Exception as e:  # noqa: BLE001 - never crash the caller on a malformed workbook
        return ParsedTable(columns=[], rows=[], warnings=[f"Failed to open workbook: {e}"])

    if not sheets:
        return ParsedTable(columns=[], rows=[], warnings=["Workbook contains no sheets."])

    row_counts = {name: _non_empty_row_count(grid) for name, grid in sheets.items()}
    non_empty_sheets = {name: count for name, count in row_counts.items() if count > 0}

    if not non_empty_sheets:
        return ParsedTable(columns=[], rows=[], warnings=["All sheets in the workbook are empty."])

    chosen_name = max(non_empty_sheets, key=non_empty_sheets.get)
    other_sheets = [name for name in sheets if name != chosen_name]
    if len(sheets) > 1:
        warnings.append(
            f"Workbook has {len(sheets)} sheets; chose '{chosen_name}' "
            f"({non_empty_sheets[chosen_name]} data rows) as the largest. "
            f"Skipped: {', '.join(other_sheets) if other_sheets else '(none)'}."
        )

    return _grid_to_table(sheets[chosen_name], warnings)
